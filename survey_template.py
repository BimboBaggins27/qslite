"""Foreman site-survey template — dispensable .xlsx the foreman fills on site.

Two-way module:
- generate_template_xlsx() → fresh blank template bytes (the foreman downloads this)
- parse_survey_xlsx(bytes) → structured SurveyHeader + list of SurveyItem (the QS-app reads this)

Layout is intentionally flat — works in Excel desktop, Excel mobile, Google Sheets, LibreOffice.
"""
from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel, Field


# ---------- data model ----------

class SurveyHeader(BaseModel):
    job_ref: str = ""
    date: str = ""
    foreman: str = ""
    client: str = ""
    site: str = ""
    scope: str = ""


class SurveyItem(BaseModel):
    zone: str = ""
    description: str = ""
    quantity: Optional[float] = None
    unit: str = ""
    photo_refs: list[str] = Field(default_factory=list)
    notes: str = ""

    def is_blank(self) -> bool:
        return not (self.description or self.zone or self.quantity or self.notes)


class SurveyData(BaseModel):
    header: SurveyHeader
    items: list[SurveyItem]

    def referenced_photos(self) -> list[str]:
        seen: list[str] = []
        for it in self.items:
            for p in it.photo_refs:
                p = p.strip()
                if p and p not in seen:
                    seen.append(p)
        return seen


# ---------- styles ----------

LABEL_FONT = Font(bold=True, size=11)
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")
HINT_FONT = Font(italic=True, size=9, color="6B7280")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------- generator ----------

VALID_UNITS = ["m", "m2", "lin.m", "each", "no", "kg", "hr", "load", "trip"]
HEADER_FIELDS: list[tuple[str, str, str]] = [
    # (cell_label, hint, field_key)
    ("Job ref / quote no.", "Free text e.g. NPQ-2026-014", "job_ref"),
    ("Date of survey", "YYYY-MM-DD or any clear date", "date"),
    ("Foreman", "Your name", "foreman"),
    ("Client", "e.g. ATESS, Bledn Properties, Raymond", "client"),
    ("Site / location", "e.g. BUILDING 1 Midrand Business Park", "site"),
    ("Scope summary", "One line: what's being quoted", "scope"),
]

ITEM_HEADERS = ["Zone / Area", "Description (what to do)", "Qty", "Unit", "Photo refs", "Notes"]
ITEM_HINTS = [
    "e.g. Bathroom 1, Kitchen, P&G's",
    "Rough is fine — 'tile floor', 'rebuild wall'",
    "Approximate — leave blank if unknown",
    "Pick from dropdown",
    "Filenames in this folder, comma-separated",
    "Anything the office needs to know",
]


def generate_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Survey"

    widths = [22, 50, 10, 10, 28, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws["A1"] = "SITE SURVEY"
    ws["A1"].font = Font(bold=True, size=16, color="0A2540")
    ws.merge_cells("A1:F1")

    ws["A2"] = "Fill what you see on site. Office will refine and price. Save with photos in same folder."
    ws["A2"].font = HINT_FONT
    ws.merge_cells("A2:F2")

    # Header block
    row = 4
    for label, hint, _key in HEADER_FIELDS:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).alignment = LEFT
        c = ws.cell(row=row, column=2, value="")
        c.alignment = LEFT
        c.border = BOX
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row=row, column=4, value=hint).font = HINT_FONT
        ws.cell(row=row, column=4).alignment = LEFT
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        row += 1

    # Spacer
    row += 1

    # Items table header
    items_header_row = row
    for col_idx, h in enumerate(ITEM_HEADERS, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BOX
    ws.row_dimensions[row].height = 22
    row += 1

    # Hint row
    for col_idx, hint in enumerate(ITEM_HINTS, start=1):
        c = ws.cell(row=row, column=col_idx, value=hint)
        c.font = HINT_FONT
        c.alignment = LEFT
        c.border = BOX
    row += 1

    # Empty rows for foreman to fill — 60 of them
    items_start = row
    for _ in range(60):
        for col_idx in range(1, 7):
            ws.cell(row=row, column=col_idx).border = BOX
        ws.row_dimensions[row].height = 22
        row += 1

    # Data validation for Unit column (col 4) — dropdown
    dv = DataValidation(type="list", formula1=f'"{",".join(VALID_UNITS)}"', allow_blank=True)
    dv.error = "Pick a unit from the dropdown"
    dv.errorTitle = "Invalid unit"
    ws.add_data_validation(dv)
    dv.add(f"D{items_start}:D{row - 1}")

    # Freeze header rows
    ws.freeze_panes = ws.cell(row=items_header_row + 2, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- parser ----------

def parse_survey_xlsx(xlsx_bytes: bytes) -> SurveyData:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    # Find the "Items" header row by scanning for "Zone / Area" in column A
    items_header_row = None
    for r in range(1, 50):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().lower().startswith("zone"):
            items_header_row = r
            break

    # Parse header block — rows 4 to (items_header - 2)
    header_dict: dict[str, str] = {}
    for label, _hint, key in HEADER_FIELDS:
        for r in range(2, (items_header_row or 30)):
            cell_label = ws.cell(row=r, column=1).value
            if isinstance(cell_label, str) and cell_label.strip().lower() == label.lower():
                val = ws.cell(row=r, column=2).value
                header_dict[key] = ("" if val is None else str(val)).strip()
                break
    header = SurveyHeader(**header_dict)

    # Parse items — start two rows below the items header row (skip hint row)
    items: list[SurveyItem] = []
    if items_header_row:
        r = items_header_row + 2  # skip the bold header AND the hint row
        while r < ws.max_row + 1:
            zone = ws.cell(row=r, column=1).value
            description = ws.cell(row=r, column=2).value
            qty = ws.cell(row=r, column=3).value
            unit = ws.cell(row=r, column=4).value
            photos = ws.cell(row=r, column=5).value
            notes = ws.cell(row=r, column=6).value

            item = SurveyItem(
                zone=str(zone or "").strip(),
                description=str(description or "").strip(),
                quantity=(float(qty) if isinstance(qty, (int, float)) and qty != 0 else None),
                unit=str(unit or "").strip(),
                photo_refs=[p.strip() for p in str(photos or "").split(",") if p.strip()],
                notes=str(notes or "").strip(),
            )
            if not item.is_blank():
                items.append(item)
            r += 1

    return SurveyData(header=header, items=items)


# ---------- helpers for the import flow ----------

def survey_to_extra_context(data: SurveyData) -> str:
    """Render the survey as a single text block we send to Claude alongside the photos."""
    lines = ["=== FOREMAN SITE SURVEY ==="]
    if data.header.job_ref: lines.append(f"Job ref: {data.header.job_ref}")
    if data.header.date: lines.append(f"Survey date: {data.header.date}")
    if data.header.foreman: lines.append(f"Foreman: {data.header.foreman}")
    if data.header.client: lines.append(f"Client: {data.header.client}")
    if data.header.site: lines.append(f"Site: {data.header.site}")
    if data.header.scope: lines.append(f"Scope summary: {data.header.scope}")
    lines.append("")
    lines.append("=== ITEMS THE FOREMAN CAPTURED ===")
    lines.append("Refine quantities from the photos, add necessary prep/finishes/compliance,")
    lines.append("and price each line. Foremen often miss prep items and SA-code compliance.")
    lines.append("")
    for i, it in enumerate(data.items, start=1):
        bits = [f"{i}."]
        if it.zone: bits.append(f"[{it.zone}]")
        if it.description: bits.append(it.description)
        qty_unit = ""
        if it.quantity is not None:
            qty_unit = f"qty≈{it.quantity}"
        if it.unit:
            qty_unit += f" {it.unit}"
        if qty_unit:
            bits.append(f"({qty_unit.strip()})")
        if it.notes:
            bits.append(f"— {it.notes}")
        if it.photo_refs:
            bits.append(f"[photos: {', '.join(it.photo_refs)}]")
        lines.append(" ".join(bits))
    return "\n".join(lines)
