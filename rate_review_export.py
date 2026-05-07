"""Rate review export for human sign-off.

Produces a single .xlsx that lists every rate currently in active use:
- catalogue rates (data/rates.json)
- learned items (memory.sqlite → learned_items)

with two empty columns for the reviewer to write into:
- 'Reviewer says' — the rate they think it should be
- 'Notes' — why

Reviewer fills those, sends back. The 'Apply reviewer edits' import (separate
function) reads the same xlsx back and updates the catalogue accordingly.
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11, color="0A2540")
SECTION_FILL = PatternFill(start_color="EAF1F8", end_color="EAF1F8", fill_type="solid")
HINT_FONT = Font(italic=True, size=9, color="6B7280")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
ZAR = "#,##0.00"


def build_rate_review_xlsx(
    catalogue: list,            # list of Rate objects
    learned: list[dict],        # memory.list_learned_items() result
    company_name: str = "Ndlovu T Projects (Pty) Ltd",
    reviewer_name: str = "Robert",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rate review"

    widths = [13, 50, 14, 8, 14, 12, 14, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title block
    ws["A1"] = f"Rate review — {company_name}"
    ws["A1"].font = Font(bold=True, size=14, color="0A2540")
    ws.merge_cells("A1:H1")

    ws["A2"] = (
        f"Reviewer: {reviewer_name}.  Goal: scan rates in column F, write your suggested rate in column G, "
        "any notes in column H. Leave G blank if the rate looks fine.  "
        f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}."
    )
    ws["A2"].font = HINT_FONT
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 32

    # Column headers (row 4)
    headers = [
        "Code",
        "Description",
        "Trade",
        "Unit",
        "Frequency / source",
        "Current rate (ZAR)",
        f"{reviewer_name} says (ZAR)",
        "Notes",
    ]
    for c_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=c_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BOX
    ws.row_dimensions[4].height = 22

    row = 5

    def section_row(label: str):
        nonlocal row
        c = ws.cell(row=row, column=1, value=label)
        c.font = SECTION_FONT
        c.fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = SECTION_FILL
            ws.cell(row=row, column=col).border = BOX
        row += 1

    # ---- Catalogue rates ----
    section_row(f"Catalogue rates ({len(catalogue)})")
    # Group by trade
    trades_seen = sorted({r.trade for r in catalogue})
    for trade in trades_seen:
        for r in (x for x in catalogue if x.trade == trade):
            ws.cell(row=row, column=1, value=r.code).border = BOX
            ws.cell(row=row, column=2, value=r.description).alignment = LEFT
            ws.cell(row=row, column=2).border = BOX
            ws.cell(row=row, column=3, value=r.trade).border = BOX
            ws.cell(row=row, column=4, value=r.unit).alignment = CENTER
            ws.cell(row=row, column=4).border = BOX
            ws.cell(row=row, column=5, value=r.source or "catalogue").alignment = LEFT
            ws.cell(row=row, column=5).border = BOX
            c = ws.cell(row=row, column=6, value=r.rate_zar)
            c.number_format = ZAR
            c.alignment = RIGHT
            c.border = BOX
            ws.cell(row=row, column=7).border = BOX  # blank for reviewer
            ws.cell(row=row, column=8).border = BOX
            row += 1

    # spacer
    row += 1

    # ---- Learned items ----
    section_row(f"Learned items from past quotes ({len(learned)})")
    if not learned:
        ws.cell(row=row, column=1, value="(none yet — issue a few quotes and they'll appear here)").font = HINT_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1
    else:
        for li in learned:
            ws.cell(row=row, column=1, value="(learned)").font = HINT_FONT
            ws.cell(row=row, column=1).border = BOX
            desc = li.get("description") or ""
            sub = li.get("subcategory") or ""
            full_desc = f"[{sub}] {desc}" if sub else desc
            ws.cell(row=row, column=2, value=full_desc).alignment = LEFT
            ws.cell(row=row, column=2).border = BOX
            ws.cell(row=row, column=3, value=li.get("trade")).border = BOX
            ws.cell(row=row, column=4, value=li.get("unit")).alignment = CENTER
            ws.cell(row=row, column=4).border = BOX
            freq = li.get("frequency") or 0
            ws.cell(row=row, column=5, value=f"{freq}× past use").alignment = LEFT
            ws.cell(row=row, column=5).border = BOX
            c = ws.cell(row=row, column=6, value=float(li.get("median_rate") or li.get("last_rate") or 0))
            c.number_format = ZAR
            c.alignment = RIGHT
            c.border = BOX
            ws.cell(row=row, column=7).border = BOX
            ws.cell(row=row, column=8).border = BOX
            row += 1

    # Footer hint
    row += 2
    ws.cell(row=row, column=1, value=(
        "Send the filled file back. We re-import it and the catalogue + learned items "
        "update so future quotes use your reviewed rates."
    )).font = HINT_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

    ws.freeze_panes = ws.cell(row=5, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
