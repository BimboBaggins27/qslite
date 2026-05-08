"""Excel renderer — pixel-faithful match to NPQ 7624 reference quote.

Same template, every export. Only client info + line items + totals vary.
Vendor block (NDLOVU header), conditions a-i, sign block, banking, and the
Plascon footer come from `company_profile.load_profile()` and never change
between quotes.

Layout, top to bottom:
  1. Logo top-left, NDLOVU vendor block top-right (name, address, phones,
     emails, Reg, VAT Reg)
  2. Thin rule
  3. Client block (bold name + address + Reg + VAT)
  4. Date (left)
  5. Quote No: (centered)
  6. ATTENTION: <name>
  7. RE: <subject> (bold + underlined)
  8. Items table — DESCRIPTION | UNIT | QTY | RATE | AMOUNT
     Zone headings inline as bold rows
  9. Totals — TOTAL EX VAT / ADD 15% VAT / TOTAL INC VAT (banded grey)
 10. Payment Terms
 11. Thank-you paragraph
 12. Conditions a-i (italic)
 13. PLEASE COMPLETE, SIGN AND RETURN
 14. Date / Quote No / Sign / ID No fields
 15. BANKING DETAILS
 16. Plascon preferred applicator footer
"""
from __future__ import annotations

import io
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LOGO_PATH = Path(__file__).parent / "assets" / "logo.png"
PLASCON_PATH = Path(__file__).parent / "assets" / "plascon-applicator.png"

# ----- Fonts -----
BODY = Font(name="Calibri", size=10)
BODY_BOLD = Font(name="Calibri", size=10, bold=True)
BODY_BOLD_UL = Font(name="Calibri", size=10, bold=True, underline="single")
BODY_SMALL = Font(name="Calibri", size=9)
BODY_SMALL_ITALIC = Font(name="Calibri", size=9, italic=True)
BODY_SMALL_BOLD_ITALIC = Font(name="Calibri", size=10, bold=True, italic=True)
VENDOR_NAME = Font(name="Calibri", size=12, bold=True)
VENDOR_META = Font(name="Calibri", size=9)
TABLE_HEADER = Font(name="Calibri", size=10, bold=True)
ZONE_FONT = Font(name="Calibri", size=10, bold=True)
SUBJECT_FONT = Font(name="Calibri", size=11, bold=True, underline="single")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)
QUOTE_NO_FONT = Font(name="Calibri", size=11, bold=True)

# ----- Borders -----
THIN = Side(style="thin", color="000000")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_RULE = Border(top=THIN)
BOTTOM_RULE = Border(bottom=THIN)

# ----- Fills -----
BAND_FILL = PatternFill("solid", fgColor="E8E8E8")  # light grey for total band

# ----- Alignment -----
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT_C = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Number format — Excel renders thousands per locale; SA Excel will show "22 755,00"
NUM_FMT = "#,##0.00"


def _zone_order(zones: list[str]) -> list[str]:
    pgs = [z for z in zones if z.lower() in ("p&g", "p&g's", "pgs", "p&gs", "site")]
    others = [z for z in zones if z not in pgs]
    return sorted(others) + sorted(pgs)


def _set(ws, row, col, value, *, font=None, alignment=None, border=None,
         number_format=None, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if alignment: c.alignment = alignment
    if border: c.border = border
    if number_format: c.number_format = number_format
    if fill: c.fill = fill
    return c


def _merge_set(ws, row, col_start, col_end, value, *, font=None, alignment=None,
               border=None, fill=None):
    """Write `value` into the first cell, then merge col_start..col_end on that row."""
    c = _set(ws, row, col_start, value, font=font, alignment=alignment,
             border=border, fill=fill)
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)
    return c


def issued_quote_to_xlsx(quote: dict) -> bytes:
    """Render the quote as an .xlsx that visually matches NPQ 7624."""
    # Lazy import so excel.py doesn't take a hard dep on company_profile during
    # tests / data-only consumers.
    try:
        from company_profile import load_profile
        company = load_profile()
    except Exception:
        company = {}

    header = quote.get("header") or {}
    items = quote.get("items", [])
    total_ex = quote.get("total_zar", 0.0)
    show_vat = bool(header.get("show_vat", company.get("show_vat", True)))
    vat_pct = float(header.get("vat_pct") or company.get("vat_pct") or 15.0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    # Page setup
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_options.horizontalCentered = True

    # 5 columns: A (Description) | B (Unit) | C (Qty) | D (Rate) | E (Amount)
    widths = [62, 8, 8, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ============================================================
    # 1. TOP BAND — logo left, NDLOVU vendor info right (cols D:E)
    # ============================================================
    if LOGO_PATH.exists():
        try:
            img = XLImage(str(LOGO_PATH))
            ratio = 130 / img.width
            img.width = 130
            img.height = int(img.height * ratio)
            img.anchor = "A1"
            ws.add_image(img)
        except Exception:
            pass

    # Vendor block in cells (right side) — merge B:E for centered display
    vendor_lines = [
        (company.get("company_name") or "NDLOVU T PROJECTS (PTY) LTD", VENDOR_NAME),
        (company.get("company_address") or "P.O. Box 702, Bergbron, 1719", VENDOR_META),
    ]
    phones: list[str] = []
    if company.get("company_phone_robert"):
        phones.append(company["company_phone_robert"])
    if company.get("company_phone_tina"):
        phones.append(company["company_phone_tina"])
    if phones:
        vendor_lines.append(("Tel: " + " | ".join(phones), VENDOR_META))
    elif company.get("company_contact"):
        vendor_lines.append((f"Tel: {company['company_contact']}", VENDOR_META))
    emails: list[str] = []
    if company.get("company_email_robert"):
        emails.append(company["company_email_robert"])
    if company.get("company_email_tina"):
        emails.append(company["company_email_tina"])
    if emails:
        vendor_lines.append(("Email: " + " | ".join(emails), VENDOR_META))
    if company.get("company_reg"):
        vendor_lines.append((f"Reg. No: {company['company_reg']}", VENDOR_META))
    if company.get("company_vat_reg"):
        vendor_lines.append((f"VAT Reg. No: {company['company_vat_reg']}", VENDOR_META))

    vendor_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for vline, vfont in vendor_lines:
        _merge_set(ws, row, 2, 5, vline, font=vfont, alignment=vendor_align)
        ws.row_dimensions[row].height = 16
        row += 1

    # Make sure logo gets enough vertical space (spans rows 1..6 visually)
    row = max(row, 8)

    # Thin rule row
    rule_row = row
    for c in range(1, 6):
        _set(ws, rule_row, c, "", border=BOTTOM_RULE)
    row += 2  # one blank below rule

    # ============================================================
    # 2. CLIENT BLOCK
    # ============================================================
    if header.get("client_name"):
        _set(ws, row, 1, header["client_name"], font=BODY_BOLD, alignment=LEFT)
        row += 1
    if header.get("client_address"):
        for line in str(header["client_address"]).split("\n"):
            line = line.strip()
            if line:
                _set(ws, row, 1, line, font=BODY_BOLD, alignment=LEFT)
                row += 1
    if header.get("client_reg"):
        _set(ws, row, 1, f"Reg: {header['client_reg']}", font=BODY_BOLD, alignment=LEFT)
        row += 1
    if header.get("client_vat_reg"):
        _set(ws, row, 1, f"VAT: {header['client_vat_reg']}", font=BODY_BOLD, alignment=LEFT)
        row += 1
    row += 1  # spacer

    # ============================================================
    # 3. DATE (left)
    # ============================================================
    if header.get("quote_date"):
        _set(ws, row, 1, header["quote_date"], font=BODY_BOLD, alignment=LEFT)
        row += 2

    # ============================================================
    # 4. QUOTE NO (centered, merged across A:E)
    # ============================================================
    if header.get("quote_no"):
        _merge_set(ws, row, 1, 5, f"Quote No: {header['quote_no']}",
                   font=QUOTE_NO_FONT,
                   alignment=Alignment(horizontal="center", vertical="center"))
        row += 2

    # ============================================================
    # 5. ATTENTION
    # ============================================================
    if header.get("attention"):
        _set(ws, row, 1, "ATTENTION:", font=BODY_BOLD, alignment=LEFT)
        _merge_set(ws, row, 2, 5, header["attention"], font=BODY_BOLD, alignment=LEFT_C)
        row += 2

    # ============================================================
    # 6. RE: subject (bold + underlined)
    # ============================================================
    if header.get("re_subject"):
        _merge_set(ws, row, 1, 5, f"RE: {header['re_subject']}",
                   font=SUBJECT_FONT, alignment=LEFT_C)
        row += 2

    # ============================================================
    # 7. ITEMS TABLE
    # ============================================================
    table_header_row = row
    headers_text = ["DESCRIPTION", "UNIT", "QTY", "RATE", "AMOUNT"]
    aligns = [LEFT_C, CENTER, CENTER, CENTER, CENTER]
    for col_idx, (h, a) in enumerate(zip(headers_text, aligns), start=1):
        _set(ws, row, col_idx, h, font=TABLE_HEADER, alignment=a, border=BOX)
    ws.row_dimensions[row].height = 22
    row += 1

    by_zone: "OrderedDict[str, list[dict]]" = OrderedDict()
    for li in items:
        z = (li.get("zone") or "Default").strip() or "Default"
        by_zone.setdefault(z, []).append(li)
    ordered_zones = _zone_order(list(by_zone.keys()))
    show_zone_labels = not (len(by_zone) == 1 and "Default" in by_zone)

    for zone in ordered_zones:
        zlist = by_zone[zone]
        if not zlist:
            continue
        if show_zone_labels and zone != "Default":
            _set(ws, row, 1, zone, font=ZONE_FONT, alignment=LEFT_C, border=BOX)
            for col in range(2, 6):
                _set(ws, row, col, "", border=BOX)
            row += 1
        for li in zlist:
            qty = float(li.get("quantity", 0))
            rate = float(li.get("rate_zar", 0))
            amount = round(qty * rate, 2)
            _set(ws, row, 1, li.get("description") or "",
                 font=BODY, alignment=LEFT_C, border=BOX)
            _set(ws, row, 2, li.get("unit") or "",
                 font=BODY, alignment=CENTER, border=BOX)
            _set(ws, row, 3, qty,
                 font=BODY, alignment=RIGHT, number_format=NUM_FMT, border=BOX)
            _set(ws, row, 4, rate,
                 font=BODY, alignment=RIGHT, number_format=NUM_FMT, border=BOX)
            _set(ws, row, 5, amount,
                 font=BODY, alignment=RIGHT, number_format=NUM_FMT, border=BOX)
            row += 1

    row += 1  # spacer before totals

    # ============================================================
    # 8. TOTALS — banded grey, three rows when VAT is shown
    # ============================================================
    vat_amount = round(total_ex * vat_pct / 100.0, 2)
    total_inc = round(total_ex + vat_amount, 2)

    def _total_row(label: str, amt: float, *, with_top: bool = False, with_bottom: bool = False):
        nonlocal row
        # Cells A..D — label, merged, light-grey filled, bold, top/bottom rules as needed
        b_top = TOP_RULE if with_top else None
        b_bot = BOTTOM_RULE if with_bottom else None
        if with_top and with_bottom:
            full_border = Border(top=THIN, bottom=THIN)
        elif with_top:
            full_border = TOP_RULE
        elif with_bottom:
            full_border = BOTTOM_RULE
        else:
            full_border = None
        _merge_set(ws, row, 1, 4, label, font=TOTAL_FONT,
                   alignment=Alignment(horizontal="left", vertical="center"),
                   fill=BAND_FILL, border=full_border)
        _set(ws, row, 5, amt, font=TOTAL_FONT, alignment=RIGHT,
             number_format=NUM_FMT, fill=BAND_FILL, border=full_border)
        ws.row_dimensions[row].height = 18
        row += 1

    if show_vat:
        _total_row("TOTAL QUOTATION EXCLUDING VAT", total_ex, with_top=True)
        _total_row(f"ADD {vat_pct:g}% VAT", vat_amount)
        _total_row("TOTAL QUOTATION INCLUDING VAT", total_inc, with_bottom=True)
    else:
        _total_row("TOTAL QUOTATION", total_ex, with_top=True, with_bottom=True)

    row += 1

    # ============================================================
    # 9. PAYMENT TERMS
    # ============================================================
    pterms = header.get("payment_terms") or company.get("payment_terms")
    if pterms:
        _set(ws, row, 1, "Payment Terms:", font=BODY_BOLD, alignment=LEFT)
        row += 1
        for line in str(pterms).splitlines():
            line = line.strip()
            if not line:
                continue
            _set(ws, row, 1, line, font=BODY_BOLD, alignment=LEFT)
            row += 1
        row += 1

    # ============================================================
    # 10. THANK-YOU PARAGRAPH (bold)
    # ============================================================
    if company.get("thank_you_paragraph"):
        _merge_set(ws, row, 1, 5, company["thank_you_paragraph"],
                   font=BODY_BOLD,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True))
        ws.row_dimensions[row].height = 36
        row += 2

    # ============================================================
    # 11. CONDITIONS a-i (italic)
    # ============================================================
    conditions = company.get("conditions") or []
    if conditions:
        _merge_set(ws, row, 1, 5,
                   "Our quotation is subject to your acceptance of the following conditions: -",
                   font=BODY_SMALL_BOLD_ITALIC,
                   alignment=Alignment(horizontal="left", vertical="top"))
        row += 1
        labels = ["a", "b", "c", "d", "e", "f", "g", "h", "I", "j", "k"]
        for i, cond in enumerate(conditions):
            label = labels[i] if i < len(labels) else f"{i+1}"
            _merge_set(ws, row, 1, 5, f"{label}) {cond}",
                       font=BODY_SMALL_ITALIC,
                       alignment=Alignment(horizontal="left", vertical="top", wrap_text=True))
            # Auto-height roughly proportional to text length
            ws.row_dimensions[row].height = 13 + 13 * (len(cond) // 90)
            row += 1
        row += 1

    # ============================================================
    # 12. ACCEPTANCE / SIGN BLOCK
    # ============================================================
    if company.get("acceptance_block"):
        _merge_set(ws, row, 1, 5, company["acceptance_block"],
                   font=BODY_SMALL_BOLD_ITALIC,
                   alignment=Alignment(horizontal="left", vertical="center"))
        row += 2
        for label in ("Date:", "Quote No:", "Sign:", "ID No:"):
            _set(ws, row, 1,
                 f"{label} ____________________",
                 font=BODY_SMALL_ITALIC, alignment=LEFT)
            row += 2
        row += 1

    # ============================================================
    # 13. BANKING DETAILS
    # ============================================================
    if company.get("banking_details"):
        _set(ws, row, 1, "BANKING DETAILS:", font=BODY_BOLD, alignment=LEFT)
        row += 1
        for line in str(company["banking_details"]).split("\n"):
            line = line.strip()
            if line:
                _set(ws, row, 1, line, font=BODY_BOLD, alignment=LEFT)
                row += 1
        row += 1

    # ============================================================
    # 14. PLASCON FOOTER (image if present, else text)
    # ============================================================
    plascon_year = company.get("plascon_applicator_year")
    if plascon_year:
        if PLASCON_PATH.exists():
            try:
                img = XLImage(str(PLASCON_PATH))
                ratio = 200 / img.width
                img.width = 200
                img.height = int(img.height * ratio)
                img.anchor = f"A{row}"
                ws.add_image(img)
                row += 4  # leave space for the image
            except Exception:
                _set(ws, row, 1, f"Plascon preferred applicator {plascon_year}",
                     font=BODY_BOLD, alignment=LEFT)
                row += 1
        else:
            _set(ws, row, 1, f"Plascon preferred applicator {plascon_year}",
                 font=BODY_BOLD, alignment=LEFT)
            row += 1

    # Set print area to fit a single page width
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f"{table_header_row}:{table_header_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
